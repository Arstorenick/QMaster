from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework.exceptions import PermissionDenied
from django.shortcuts import get_object_or_404
from django.db.models import Count, Q

from apps.surveys.models import Survey, Question
from .models import Submission, Answer
from .serializers import SubmissionCreateSerializer, SubmissionListSerializer


def check_owner(user, survey):
    """检查用户是否为问卷所有者"""
    if survey.owner != user:
        raise PermissionDenied('无权操作此问卷')


# ══════════════════════════════════════════════════════════
#  Submit (public)
# ══════════════════════════════════════════════════════════

@api_view(['POST'])
@permission_classes([AllowAny])
def submit_survey(request, survey_id):
    """提交问卷答案（公开）。每个用户仅可提交一次"""
    survey = get_object_or_404(Survey, id=survey_id, status=1, is_deleted=False)

    # 检查是否已提交（按 IP + 问卷判断，登录用户按 user 判断）
    if request.user.is_authenticated:
        existing = Submission.objects.filter(survey=survey, submit_ip=request.META.get('REMOTE_ADDR', '')).exists()
        # 更严格：按用户+问卷
        existing = Submission.objects.filter(
            survey=survey,
            submit_ip__in=[request.META.get('REMOTE_ADDR', ''), request.META.get('HTTP_X_FORWARDED_FOR', '')]
        ).exists()
        if existing:
            return Response({'detail': '您已提交过此问卷，不可重复提交'}, status=status.HTTP_400_BAD_REQUEST)

    # 验证必填题目（排除分页、分段、分割线等不可答题类型）
    non_answerable = ['page_break', 'section_break', 'divider', 'image_carousel']
    required_questions = set(
        survey.questions.filter(is_required=True).exclude(type__in=non_answerable).values_list('id', flat=True)
    )
    submitted_question_ids = set()
    for ans in request.data.get('answers', []):
        qid = ans.get('question')
        has_value = any([
            ans.get('option'),
            ans.get('answer_text'),
            ans.get('answer_number'),
            ans.get('answer_file'),
            ans.get('answer_json'),
        ])
        if qid and has_value:
            submitted_question_ids.add(qid)

    missing = required_questions - submitted_question_ids
    if missing:
        missing_titles = list(
            Question.objects.filter(id__in=missing).values_list('title', flat=True)
        )
        return Response(
            {'detail': '请完成所有必填题目', 'missing_questions': missing_titles},
            status=status.HTTP_400_BAD_REQUEST
        )

    serializer = SubmissionCreateSerializer(
        data=request.data, context={'request': request}
    )
    if serializer.is_valid():
        submission = serializer.save(survey=survey)
        return Response(
            SubmissionListSerializer(submission).data,
            status=status.HTTP_201_CREATED
        )
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# ══════════════════════════════════════════════════════════
#  Statistics (owner only)
# ══════════════════════════════════════════════════════════

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def survey_statistics(request, survey_id):
    """问卷数据统计"""
    survey = get_object_or_404(Survey, id=survey_id, is_deleted=False)
    check_owner(request.user, survey)

    total_submissions = survey.submissions.count()
    questions = survey.questions.exclude(
        type__in=['page_break', 'section_break', 'divider', 'image_carousel']
    ).order_by('order')

    results = []
    for q in questions:
        q_data = {
            'question_id': q.id,
            'title': q.title,
            'type': q.type,
            'total_answers': Answer.objects.filter(
                question=q, submission__survey=survey
            ).count(),
        }

        if q.type in ('radio', 'dropdown', 'image_radio'):
            q_data['options'] = _calc_option_stats(q, total_submissions)
        elif q.type == 'checkbox':
            q_data['options'] = _calc_option_stats(q, total_submissions)
        elif q.type in ('rating', 'scale', 'slider', 'matrix_numeric'):
            q_data['avg'] = _calc_avg(q, survey)
        elif q.type in ('text', 'textarea'):
            q_data['text_count'] = Answer.objects.filter(
                question=q, submission__survey=survey
            ).exclude(answer_text='').count()

        results.append(q_data)

    return Response({
        'survey_title': survey.title,
        'total_submissions': total_submissions,
        'questions': results,
    })


def _calc_option_stats(question, total_submissions):
    """计算各选项的选择数量和百分比"""
    stats = []
    for opt in question.options.all().order_by('order'):
        count = Answer.objects.filter(
            question=question, option=opt, submission__survey=question.survey
        ).count()
        stats.append({
            'option_id': opt.id,
            'title': opt.title,
            'count': count,
            'percentage': round(count / total_submissions * 100, 1) if total_submissions > 0 else 0,
        })
    return stats


def _calc_avg(question, survey):
    """计算数值题平均分"""
    from django.db.models import Avg
    result = Answer.objects.filter(
        question=question, submission__survey=survey
    ).aggregate(avg=Avg('answer_number'))
    return round(result['avg'], 2) if result['avg'] else 0


# ══════════════════════════════════════════════════════════
#  Text Answers
# ══════════════════════════════════════════════════════════

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def text_answers(request, survey_id, question_id):
    """获取文本题答案列表（分页）"""
    survey = get_object_or_404(Survey, id=survey_id, is_deleted=False)
    check_owner(request.user, survey)
    question = get_object_or_404(Question, id=question_id, survey=survey)

    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 20))
    offset = (page - 1) * page_size

    answers = Answer.objects.filter(
        question=question, submission__survey=survey
    ).exclude(answer_text='').select_related('submission').order_by(
        '-submission__submit_time'
    )[offset:offset + page_size]

    total = Answer.objects.filter(
        question=question, submission__survey=survey
    ).exclude(answer_text='').count()

    return Response({
        'total': total,
        'page': page,
        'page_size': page_size,
        'answers': [
            {
                'answer_text': a.answer_text,
                'submit_time': a.submission.submit_time,
            } for a in answers
        ],
    })


# ══════════════════════════════════════════════════════════
#  Export (Excel)
# ══════════════════════════════════════════════════════════

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_statistics(request, survey_id):
    """导出统计结果为 Excel"""
    import io
    from openpyxl import Workbook

    survey = get_object_or_404(Survey, id=survey_id, is_deleted=False)
    check_owner(request.user, survey)

    wb = Workbook()
    ws = wb.active
    ws.title = '数据统计'

    questions = survey.questions.exclude(
        type__in=['page_break', 'section_break', 'divider', 'image_carousel']
    ).order_by('order')

    total_submissions = survey.submissions.count()

    row = 1
    ws.cell(row=row, column=1, value=f'问卷: {survey.title}')
    ws.cell(row=row, column=2, value=f'总提交数: {total_submissions}')
    row += 2

    for q in questions:
        ws.cell(row=row, column=1, value=q.title)
        ws.cell(row=row, column=2, value=f'[题型: {q.get_type_display()}]')
        row += 1

        if q.type in ('radio', 'checkbox', 'dropdown', 'image_radio', 'image_checkbox'):
            ws.cell(row=row, column=1, value='选项')
            ws.cell(row=row, column=2, value='数量')
            ws.cell(row=row, column=3, value='占比')
            row += 1
            for opt in q.options.all():
                count = Answer.objects.filter(
                    question=q, option=opt, submission__survey=survey
                ).count()
                ws.cell(row=row, column=1, value=opt.title)
                ws.cell(row=row, column=2, value=count)
                pct = round(count / total_submissions * 100, 1) if total_submissions > 0 else 0
                ws.cell(row=row, column=3, value=f'{pct}%')
                row += 1

        elif q.type in ('rating', 'scale', 'slider'):
            avg = _calc_avg(q, survey)
            ws.cell(row=row, column=1, value=f'平均分: {avg}')

        row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from django.http import HttpResponse
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{survey.title}_统计.xlsx"'
    return response
